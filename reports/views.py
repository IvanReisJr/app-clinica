from django.http import HttpResponse
from django.template.loader import get_template
from django.utils import timezone
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.exceptions import PermissionDenied
from patients.models import Patient
from appointments.models import Appointment
from medications.models import Medication, MedicationMovement
from users.models import RolePermission
from system_settings.models import SystemSetting
from xhtml2pdf import pisa
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

def _get_clinic_context():
    settings = {s.key: s.value for s in SystemSetting.objects.all()}
    return {
        'clinic_name': settings.get('clinic_name', 'MEDTRACE'),
        'clinic_logo': settings.get('clinic_logo_url', ''),
    }

class ExportPatientsView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if request.user.role != 'admin':
            if not RolePermission.objects.filter(role=request.user.role, permission_slug='view_reports', is_granted=True).exists():
                raise PermissionDenied("Você não tem permissão para visualizar relatórios.")

        export_type = request.query_params.get('export_type', request.query_params.get('format', 'pdf')).lower()
        patients = Patient.objects.filter(is_active=True).order_by('full_name')

        if export_type == 'excel':
            return self.export_excel(patients)
        else:
            return self.export_pdf(patients)

    def export_pdf(self, patients):
        template = get_template('reports/patients_report.html')
        context = {
            'title': 'Relatório Geral de Pacientes',
            'patients': patients,
            'now': timezone.now(),
            'user': self.request.user,
            **_get_clinic_context()
        }
        html = template.render(context)
        
        result = BytesIO()
        pdf = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), result)
        
        if not pdf.err:
            response = HttpResponse(result.getvalue(), content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="relatorio_pacientes_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
            return response
        
        return HttpResponse("Erro ao gerar PDF", status=400)

    def export_excel(self, patients):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Pacientes"

        # Cabeçalho
        columns = ['Nome Completo', 'CPF', 'Data de Nascimento', 'Gênero', 'Telefone', 'E-mail', 'Convênio']
        ws.append(columns)

        # Estilizar cabeçalho
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1e293b", end_color="1e293b", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Dados
        for p in patients:
            ws.append([
                p.full_name,
                p.cpf,
                p.date_of_birth.strftime('%d/%m/%Y') if p.date_of_birth else '',
                p.get_gender_display() if p.gender else '',
                p.phone or '',
                p.email or '',
                p.insurance or 'Particular'
            ])

        # Ajustar largura das colunas
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="relatorio_pacientes_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        return response

class ExportAppointmentsView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if request.user.role != 'admin':
            if not RolePermission.objects.filter(role=request.user.role, permission_slug='view_reports', is_granted=True).exists():
                raise PermissionDenied("Você não tem permissão para visualizar relatórios.")

        export_type = request.query_params.get('export_type', request.query_params.get('format', 'pdf')).lower()
        # Filtro simples: hoje ou todos. Para um MVP, vamos pegar todos ativos
        appointments = Appointment.objects.filter(is_active=True).select_related('patient', 'professional').order_by('appointment_date', 'appointment_time')

        if export_type == 'excel':
            return self.export_excel(appointments)
        else:
            return self.export_pdf(appointments)

    def export_pdf(self, appointments):
        template = get_template('reports/appointments_report.html')
        context = {
            'title': 'Relatório de Agendamentos',
            'appointments': appointments,
            'now': timezone.now(),
            'user': self.request.user,
            **_get_clinic_context()
        }
        html = template.render(context)
        
        result = BytesIO()
        pdf = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), result)
        
        if not pdf.err:
            response = HttpResponse(result.getvalue(), content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="relatorio_agendamentos_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
            return response
        return HttpResponse("Erro ao gerar PDF", status=400)

    def export_excel(self, appointments):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Agendamentos"

        # Cabeçalho
        columns = ['Data', 'Hora', 'Paciente', 'Profissional', 'Status', 'Encaixe']
        ws.append(columns)

        # Estilizar cabeçalho
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1e293b", end_color="1e293b", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Dados
        for a in appointments:
            ws.append([
                a.appointment_date.strftime('%d/%m/%Y'),
                a.appointment_time.strftime('%H:%M'),
                a.patient.full_name,
                a.professional.name if a.professional else '-',
                a.get_status_display(),
                'Sim' if a.is_encaixe else 'Não'
            ])

        # Ajustar largura das colunas
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except: pass
            ws.column_dimensions[column].width = (max_length + 2)

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="relatorio_agendamentos_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        return response

class ExportInventoryView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if request.user.role != 'admin':
            if not RolePermission.objects.filter(role=request.user.role, permission_slug='view_reports', is_granted=True).exists():
                raise PermissionDenied("Você não tem permissão para visualizar relatórios.")

        export_type = request.query_params.get('export_type', request.query_params.get('format', 'pdf')).lower()
        # Filtro: Itens com pouco estoque ou todos
        report_id = request.query_params.get('report_id', 'current_inventory')
        
        medications = Medication.objects.filter(is_active=True).order_by('name')
        if report_id == 'low_stock':
            # Consideramos baixo estoque itens com < 10 unidades por simplicidade agora
            medications = medications.filter(quantity__lt=10)

        if export_type == 'excel':
            return self.export_excel(medications, report_id)
        else:
            return self.export_pdf(medications, report_id)

    def export_pdf(self, medications, report_id):
        template = get_template('reports/inventory_report.html')
        title = "Alertas de Reposição (Estoque Baixo)" if report_id == 'low_stock' else "Inventário Atual de Medicamentos"
        context = {
            'title': title,
            'medications': medications,
            'now': timezone.now(),
            'user': self.request.user,
            **_get_clinic_context()
        }
        html = template.render(context)
        result = BytesIO()
        pdf = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), result)
        if not pdf.err:
            response = HttpResponse(result.getvalue(), content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="inventario_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
            return response
        return HttpResponse("Erro ao gerar PDF", status=400)

    def export_excel(self, medications, report_id):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Estoque"
        columns = ['Nome do Medicamento', 'Lote', 'Validade', 'Qtd em Estoque', 'Fornecedor']
        ws.append(columns)

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1e293b", end_color="1e293b", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for m in medications:
            ws.append([
                m.name,
                m.lot_number or '-',
                m.expiration_date.strftime('%d/%m/%Y') if m.expiration_date else '-',
                m.quantity,
                m.provider or '-'
            ])

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except: pass
            ws.column_dimensions[column].width = (max_length + 2)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="inventario_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        return response

class ExportMedicationMovementsView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if request.user.role != 'admin':
            if not RolePermission.objects.filter(role=request.user.role, permission_slug='view_reports', is_granted=True).exists():
                raise PermissionDenied("Você não tem permissão para visualizar relatórios.")

        export_type = request.query_params.get('export_type', request.query_params.get('format', 'pdf')).lower()
        movements = MedicationMovement.objects.all().select_related('medication', 'user').order_by('-created_at')

        if export_type == 'excel':
            return self.export_excel(movements)
        else:
            # Injeta a quantidade formatada para evitar quebra de linha no template
            for m in movements:
                prefix = "+" if m.type == 'entrada' else "-" if m.type in ['saida', 'vencimento'] else ""
                m.formatted_quantity = f"{prefix}{m.quantity}"
            return self.export_pdf(movements)

    def export_pdf(self, movements):
        template = get_template('reports/movements_report.html')
        context = {
            'title': 'Histórico de Movimentação de Estoque (Kardex)',
            'movements': movements,
            'now': timezone.now(),
            'user': self.request.user,
            **_get_clinic_context()
        }
        html = template.render(context)
        result = BytesIO()
        pdf = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), result)
        if not pdf.err:
            response = HttpResponse(result.getvalue(), content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="kardex_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
            return response
        return HttpResponse("Erro ao gerar PDF", status=400)

    def export_excel(self, movements):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Movimentacoes"
        columns = ['Data/Hora', 'Medicamento', 'Tipo', 'Quantidade', 'Usuário', 'Motivo']
        ws.append(columns)

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1e293b", end_color="1e293b", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for m in movements:
            ws.append([
                m.created_at.strftime('%d/%m/%Y %H:%M'),
                m.medication.name,
                m.get_type_display(),
                m.quantity,
                m.user.username if m.user else '-',
                m.description or '-'
            ])

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except: pass
            ws.column_dimensions[column].width = (max_length + 2)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="kardex_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        return response
